from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
import openpyxl


PROFILE_PATH = r'C:\Users\apfri\AppData\Local\Google\Chrome\User Data\Default'

options = webdriver.ChromeOptions()
options.add_argument('--start-maximized')
options.add_argument(f'user-data-dir={PROFILE_PATH}')
driver = webdriver.Chrome(options=options)

url = 'https://www.terabyteshop.com.br/pc-gamer/t-gamer'

driver.get(url)
prod_name_element = '//a[@class="prod-name"]'
prod_new_price_element = '//div[@class = "prod-new-price"]/span'

# workbook = openpyxl.Workbook()
# workbook.create_sheet('Produtos')
# sheet_produtos = workbook['Produtos']
# sheet_produtos['A1'] = 'Produto'
# sheet_produtos['B1'] = 'Preço'
# sleep(3)
# prod_name = driver.find_elements(By.XPATH, prod_name_element)
# prod_price = driver.find_elements(By.XPATH, prod_new_price_element)
# i = 2
# for prod , price in zip(prod_name, prod_price):
#     sheet_produtos[f'A{i}'] = prod.text
#     i += 1

# workbook.save('terabyte.xlsx')


workbook = openpyxl.load_workbook('terabyte.xlsx')
sheet_produtos = workbook['Produtos']

search_element = '//input[@id="isearch"]'
for i, rows in enumerate(sheet_produtos.iter_rows(min_row=2, min_col=1,max_col=2, values_only=True),start=2):
    if rows[0] and not rows[1]:
        sleep(1)
        search = driver.find_element(By.XPATH, search_element)
        search.send_keys(rows[0])
        search.submit()
        sleep(1)
        try:
            new_price = driver.find_element(By.XPATH, prod_new_price_element).text
        except:
            new_price = 'Sem Estoque'
        print(new_price)
        sheet_produtos.cell(row=i, column=2, value=new_price)
        workbook.save('terabyte.xlsx')