from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
import openpyxl

product_element = '//input[@id = "product_name"]'
description_element = '//textarea[@id = "description"]'
category_element = '//input[@id = "category"]'
product_code_element = '//input[@id = "product_code"]'
weight_element = '//input[@id = "weight"]'
dimensions_element = '//input[@id = "dimensions"]'
button_next_element =  '//button[contains(concat(" ", @class, " ")," btn-primary ")]'
price_element = '//input[@id = "price"]'
stock_element = '//input[@id = "stock"]'
expiry_date_element = '//input[@id = "expiry_date"]'
color_element = '//input[@id = "color"]'
size_element = '//select[@id = "size"]'
material_element = '//input[@id = "material"]'
# Clicar em próximo (button_next_element)
manufacturer_element = '//input[@id = "manufacturer"]'
country_element = '//input[@id = "country"]'
remarks_element = '//textarea[@id = "remarks"]'
bar_code_element = '//input[@id = "barcode"]'
warehouse_location_element = '//input[@id = "warehouse_location"]'
# Clicar em Concluir (button_next_element)
# Clicar em Adicionar mais um (button_next_element)


options = webdriver.ChromeOptions()
options.add_argument('--start-maximized')
driver = webdriver.Chrome(options=options)

driver.get('https://cadastro-produtos-devaprender.netlify.app/')
sleep(2)

def fxpth(element):
    return driver.find_element(By.XPATH, element)


elementos = [
    product_element,
    description_element,
    category_element,
    product_code_element,
    weight_element,
    dimensions_element,
    price_element,
    stock_element,
    expiry_date_element,
    color_element,
    size_element,
    material_element,
    manufacturer_element,
    country_element,
    remarks_element,
    bar_code_element,
    warehouse_location_element
]
workbook = openpyxl.load_workbook('dev-aprender/produtos_ficticios.xlsx')
produtos_sheet = workbook['Produtos']
for row in produtos_sheet.iter_rows(min_row=2):
    if row:
        for i, cell in enumerate(row):
            if i == 6 or i == 12:
                botao = fxpth(button_next_element)
                botao.click()
                sleep(1)
            element = fxpth(elementos[i])
            element.send_keys(cell.value)

        proximo = fxpth(button_next_element)
        proximo.click()
        sleep(.5)
        alerta = driver.switch_to.alert
        alerta.accept()
        sleep(2)
        proximo = fxpth(button_next_element)
        proximo.click()
        sleep(2)

print(workbook.sheetnames)